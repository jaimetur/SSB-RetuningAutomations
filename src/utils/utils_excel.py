# -*- coding: utf-8 -*-

import re
from typing import Set

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


# ============================ EXCEL HELPERS ============================

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[:\\/?*\[\]]', "_", name)
    name = name.strip().strip("'")
    return (name or "Sheet")[:31]


def unique_sheet_name(base: str, used: Set[str]) -> str:
    if base not in used:
        return base
    for k in range(1, 1000):
        suffix = f" ({k})"
        cand = (base[: max(0, 31 - len(suffix))] + suffix)
        if cand not in used:
            return cand
    i, cand = 1, base
    while cand in used:
        cand = f"{base[:28]}_{i:02d}"
        i += 1
    return cand

def apply_alternating_category_row_fills(
    ws: Worksheet,
    category_header: str = "Category",
    header_row: int = 1,
    start_row: int | None = None,
    end_row: int | None = None,
    fill_color_1: str = "E0F7FA",
    fill_color_2: str = "B2EBF2",
    value_header: str = "Value",
) -> None:
    """
    Apply alternating background fills to row blocks based on Category changes.

    Each time the value in the Category column changes, the row fill color
    toggles between two similar colors so that each Category block is visually
    separated in the Excel sheet.

    Additionally:
    - Any row whose SubCategory contains the string "inconsist" (case-insensitive)
      will have its font colored red if Value > 0, or dark gray otherwise.
    """

    # Find the Category column index based on the header name
    category_col_idx: int | None = None
    subcategory_col_idx: int | None = None
    value_col_idx: int | None = None

    value_header_norm = str(value_header).strip().lower()

    for cell in ws[header_row]:
        header_value = str(cell.value).strip() if cell.value is not None else ""
        header_lower = header_value.lower()

        if header_value == category_header:
            category_col_idx = cell.column
        elif header_lower == "subcategory":
            subcategory_col_idx = cell.column
        elif header_lower == value_header_norm:
            value_col_idx = cell.column

    if category_col_idx is None:
        # Category column not found, nothing to do
        return

    if start_row is None:
        start_row = header_row + 1
    if end_row is None:
        end_row = ws.max_row

    fill1 = PatternFill(fill_type="solid", fgColor=fill_color_1)
    fill2 = PatternFill(fill_type="solid", fgColor=fill_color_2)

    # Fonts for inconsistency rows
    red_font = Font(color="FF0000")      # Value > 0
    gray_font = Font(color="A6A6A6")     # Value <= 0 or non-numeric

    last_category = None
    use_first = True
    current_fill = fill1

    for row_idx in range(start_row, end_row + 1):

        # Read the Category value in the current row
        cell_category = ws.cell(row=row_idx, column=category_col_idx).value

        # Detect category changes → toggle background color
        if cell_category != last_category:
            if last_category is None:
                use_first = True
            else:
                use_first = not use_first
            current_fill = fill1 if use_first else fill2
            last_category = cell_category

        # Detect whether this row is an "Inconsistencies" or "Discrepancy" row
        is_inconsistency_or_discrepancy_row = False
        if subcategory_col_idx is not None:
            sub_val = ws.cell(row=row_idx, column=subcategory_col_idx).value
            if sub_val is not None:
                sub_norm = str(sub_val).strip().lower()
                if "inconsist" in sub_norm or "discrep" in sub_norm:
                    is_inconsistency_or_discrepancy_row = True

        # Determine if the inconsistency or discrepancy has Value > 0
        is_positive_inconsistency_or_discrepancy = False
        if is_inconsistency_or_discrepancy_row and value_col_idx is not None:
            value_cell = ws.cell(row=row_idx, column=value_col_idx).value
            num_val = None
            if isinstance(value_cell, (int, float)):
                num_val = float(value_cell)
            else:
                try:
                    if value_cell is not None:
                        num_val = float(str(value_cell).strip())
                except Exception:
                    num_val = None
            if num_val is not None and num_val > 0:
                is_positive_inconsistency_or_discrepancy = True

        # Apply background color and (optionally) red/gray font to the entire row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = current_fill
            if is_inconsistency_or_discrepancy_row:
                cell.font = red_font if is_positive_inconsistency_or_discrepancy else gray_font



def color_summary_tabs(writer, prefix: str = "Summary", rgb_hex: str = "00B050") -> None:
    """
    Set tab color for every worksheet whose name starts with `prefix`.
    Works with openpyxl-backed ExcelWriter.
    - rgb_hex: 6-hex RGB (e.g., '00B050' = green).
    """
    try:
        wb = writer.book  # openpyxl Workbook
        for ws in wb.worksheets:
            if ws.title.startswith(prefix):
                # Set tab color (expects hex without '#')
                ws.sheet_properties.tabColor = rgb_hex
    except Exception:
        # Hard-fail safe: never break file writing just for coloring tabs
        pass


def enable_header_filters(writer, freeze_header: bool = True, align_left: bool = True) -> None:
    """
    Enable Excel AutoFilter on every worksheet for the used range.
    Optionally freeze the header row (row 1) so data scrolls under it.
    Optionally align all cell text to the left.
    """
    try:
        wb = writer.book  # openpyxl Workbook
        for ws in wb.worksheets:
            # Skip empty sheets safely
            if ws.max_row < 1 or ws.max_column < 1:
                continue

            # Define used range for the filter, from A1 to last used cell
            top_left = ws.cell(row=1, column=1).coordinate
            bottom_right = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
            ws.auto_filter.ref = f"{top_left}:{bottom_right}"

            # Optionally freeze header row
            if freeze_header and ws.max_row >= 2:
                ws.freeze_panes = "A2"

            # Optionally align all text to the left
            if align_left:
                left_alignment = Alignment(horizontal="left", vertical="top")
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None:
                            cell.alignment = left_alignment

    except Exception:
        # Never fail the export just for filters or alignment
        pass


def style_headers_autofilter_and_autofit(writer, freeze_header: bool = True, align: str = "left", header_color: str = "CCE5FF", max_width: int = 100, autofit_rows: object = 50, enable_a1_hyperlink: bool = False, hyperlink_sheet: str = "SummaryAudit", category_sheet_map: dict | None = None, summary_category_header: str = "Category", max_summary_category_links: int = 800) -> None:
    """
    Apply header styling (with configurable color), enable auto-filter on the first row,
    freeze header optionally, and auto-fit all column widths.

    Extra (optional) features:
      - A1 internal hyperlink to a summary sheet (e.g. SummaryAudit or Summary_CellRelation)
      - Hyperlinks inside the summary sheet: Summary.Category -> target sheet (A1)

    Params:
        writer: ExcelWriter object (pandas)
        freeze_header: whether to freeze the first row (default=True)
        align: horizontal alignment for header text ("left", "center", "right")
        header_color: fill color for header row (hex string, default="CCE5FF")
        max_width: maximum allowed column width when auto-fitting (default=50)
        autofit_rows: number of content rows (starting from row 2) to sample when auto-fitting.
                     - default: 50
                     - "All": scan all rows
                     - any int <= 0: behaves like 0 (header-only)
        a1_hyperlink_to_summaryaudit: if True, set A1 hyperlink in every sheet (except summaryaudit_sheet)
        summaryaudit_sheet: sheet name acting as the summary sheet (default="SummaryAudit")
        category_sheet_map: optional mapping to resolve Category values to actual sheet names
        summary_category_header: header name to use for hyperlinking rows in the summary sheet (default="Category")
        max_summary_category_links: safety limit for how many summary rows will be hyperlinked (default=800)
    """
    workbook = writer.book

    # Normalize autofit_rows
    use_all_rows = False
    rows_to_scan = 50
    if isinstance(autofit_rows, str) and autofit_rows.strip().lower() == "all":
        use_all_rows = True
    else:
        try:
            rows_to_scan = int(autofit_rows)
        except Exception:
            rows_to_scan = 50  # safe default
        if rows_to_scan < 0:
            rows_to_scan = 0

    sheetnames = []
    try:
        sheetnames = list(getattr(workbook, "sheetnames", []) or [])
    except Exception:
        sheetnames = []

    if not sheetnames:
        try:
            sheetnames = [getattr(ws, "title", "") for ws in getattr(workbook, "worksheets", [])]
        except Exception:
            sheetnames = []

    has_summary_sheet = bool(hyperlink_sheet) and (hyperlink_sheet in sheetnames)

    for ws in getattr(workbook, "worksheets", []):
        # Skip sheets without content
        try:
            if ws.max_row < 1 or ws.max_column < 1:
                continue
        except Exception:
            continue

        # --------------------------------------------------------------
        # 1) Apply header style
        # --------------------------------------------------------------
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        header_alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

        try:
            for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = Font(bold=True, color="000000")
                    cell.alignment = header_alignment
        except Exception:
            pass

        # --------------------------------------------------------------
        # 2) Apply auto-filter on the full table
        # --------------------------------------------------------------
        try:
            first_col = get_column_letter(1)
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"{first_col}1:{last_col}{ws.max_row}"
        except Exception:
            pass

        # --------------------------------------------------------------
        # 3) Freeze header row if enabled
        # --------------------------------------------------------------
        if freeze_header:
            try:
                ws.freeze_panes = "A2"
            except Exception:
                pass

        # --------------------------------------------------------------
        # 4) Auto-fit column widths (sample first N rows by default)
        # --------------------------------------------------------------
        try:
            if use_all_rows:
                row_end = ws.max_row
            else:
                row_end = min(ws.max_row, 1 + rows_to_scan)

            max_lens = [0] * ws.max_column
            for row in ws.iter_rows(min_row=1, max_row=row_end, min_col=1, max_col=ws.max_column, values_only=True):
                for j, v in enumerate(row):
                    try:
                        s = "" if v is None else str(v)
                        s = s.replace("\r\n", " ").replace("\n", " ").strip()
                        l = len(s)
                        if l > max_lens[j]:
                            max_lens[j] = l
                    except Exception:
                        pass

            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max_lens[col_idx - 1] + 2, max_width)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # 5) Optional hyperlinks (OpenPyXL engine)
    # ------------------------------------------------------------------
    if enable_a1_hyperlink and has_summary_sheet:
        for ws in getattr(workbook, "worksheets", []):
            try:
                if getattr(ws, "title", "") == hyperlink_sheet:
                    continue
                cell = ws.cell(row=1, column=1)
                cell.hyperlink = f"#'{hyperlink_sheet}'!A1"
                try:
                    cell.font = cell.font.copy(color="0563C1", underline="single")
                except Exception:
                    try:
                        cell.font = Font(name=cell.font.name, sz=cell.font.sz, bold=cell.font.bold, italic=cell.font.italic, vertAlign=cell.font.vertAlign, underline="single", strike=cell.font.strike, color="0563C1")
                    except Exception:
                        cell.font = Font(color="0563C1", underline="single")
            except Exception:
                pass

    if has_summary_sheet and bool(summary_category_header):
        try:
            ws_summary = workbook[hyperlink_sheet]
        except Exception:
            ws_summary = None

        if ws_summary is not None:
            try:
                header_vals = []
                for row in ws_summary.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws_summary.max_column, values_only=True):
                    header_vals = [str(v).strip() if v is not None else "" for v in row]

                try:
                    category_col_idx = header_vals.index(str(summary_category_header)) + 1
                except ValueError:
                    category_col_idx = None

                if category_col_idx:
                    try:
                        max_rows = int(max_summary_category_links)
                    except Exception:
                        max_rows = 800
                    if max_rows < 0:
                        max_rows = 0

                    last_row = ws_summary.max_row
                    if max_rows and last_row > (1 + max_rows):
                        last_row = 1 + max_rows

                    for r in range(2, last_row + 1):
                        try:
                            cell = ws_summary.cell(row=r, column=category_col_idx)
                            raw = str(cell.value).strip() if cell.value is not None else ""
                            if not raw:
                                continue

                            target_sheet = raw
                            if category_sheet_map and target_sheet and target_sheet not in sheetnames:
                                target_sheet = str(category_sheet_map.get(raw, raw))

                            if target_sheet and target_sheet in sheetnames:
                                cell.hyperlink = f"#'{target_sheet}'!A1"
                                try:
                                    cell.font = cell.font.copy(color="0563C1", underline="single")
                                except Exception:
                                    cell.font = Font(color="0563C1", underline="single")
                        except Exception:
                            pass
            except Exception:
                pass


